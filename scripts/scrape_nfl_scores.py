#!/usr/bin/env python3
"""
NFL Historical Scores Scraper
Scrapes NFL game scores from Pro Football Reference and exports to Excel.
Supports CLI arguments for year range and output path.
"""

import argparse
import subprocess
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
import sys
import os
import re

# ── Defaults ──
DEFAULT_START_YEAR = 1966
DEFAULT_END_YEAR = 2024
BASE_URL = "https://www.pro-football-reference.com/years/{year}/games.htm"
REQUEST_DELAY = 3.5
MAX_RETRIES = 3

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUTPUT = os.path.join(SCRIPT_DIR, "nfl_scores.xlsx")

USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

COLUMNS = [
    "Season", "Week", "Day", "Date", "Time (ET)",
    "Home Team", "Away Team",
    "Home Score", "Away Score", "Score Difference",
    "Winner", "Primetime Slot"
]


def fetch_page(url, retries=MAX_RETRIES):
    """Fetch a URL using curl (bypasses PFR's Python request blocking)."""
    for attempt in range(retries):
        try:
            result = subprocess.run(
                [
                    "curl", "-s", "-L",
                    "-A", USER_AGENT,
                    "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9",
                    "-H", "Accept-Language: en-US,en;q=0.9",
                    "-H", "Accept-Encoding: identity",
                    "--max-time", "30",
                    "-w", "\n%{http_code}",
                    url
                ],
                capture_output=True, text=True, timeout=45
            )
            output = result.stdout
            lines = output.rsplit("\n", 1)
            if len(lines) == 2:
                html, status_code = lines[0], lines[1].strip()
            else:
                html, status_code = output, "000"

            if status_code == "200" and html:
                return html
            elif status_code == "429":
                wait = (attempt + 1) * 15
                print(f"  Rate limited (429). Waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  HTTP {status_code} for {url}")
                time.sleep(5)
        except (subprocess.TimeoutExpired, Exception) as e:
            print(f"  Fetch error (attempt {attempt+1}): {e}")
            time.sleep(5)
    return None


def parse_time(time_str):
    """Parse time string like '8:20PM' into (hour_24, minute)."""
    if not time_str:
        return None, None
    time_str = time_str.strip().upper()
    match = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str)
    if not match:
        return None, None
    hour = int(match.group(1))
    minute = int(match.group(2))
    ampm = match.group(3)
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    return hour, minute


def detect_primetime(day, time_str, season):
    """Detect if a game is a primetime game."""
    if not day or not time_str:
        return ""
    day = day.strip()
    hour, minute = parse_time(time_str)

    if day == "Mon" and season >= 1970:
        if hour is not None and hour >= 19:
            return "MNF"
        return "MNF"

    if day == "Thu":
        if season >= 2006 and hour is not None and hour >= 19:
            return "TNF"
        elif season >= 2006:
            return "TNF"
        return ""

    if day == "Sun":
        if hour is not None and hour >= 19:
            if season >= 1987:
                return "SNF"
        return ""

    if day == "Sat":
        if hour is not None and hour >= 19:
            return "Saturday Primetime"
        return ""

    return ""


def scrape_season(year):
    """Scrape all games for a given season from PFR."""
    url = BASE_URL.format(year=year)
    html = fetch_page(url)
    if not html:
        print(f"  FAILED to fetch {year}")
        return []

    soup = BeautifulSoup(html, "html.parser")

    table = soup.find("table", id="games")
    if not table:
        table = soup.find("table")
        if not table:
            print(f"  No table found for {year}")
            return []

    tbody = table.find("tbody")
    if not tbody:
        print(f"  No tbody found for {year}")
        return []

    games = []
    for row in tbody.find_all("tr"):
        if row.find("th", {"scope": "col"}):
            continue

        week_el = row.find(["th", "td"], {"data-stat": "week_num"})
        day_el = row.find("td", {"data-stat": "game_day_of_week"})
        date_el = row.find("td", {"data-stat": "game_date"})
        time_el = row.find("td", {"data-stat": "gametime"})
        winner_el = row.find("td", {"data-stat": "winner"})
        location_el = row.find("td", {"data-stat": "game_location"})
        loser_el = row.find("td", {"data-stat": "loser"})
        pts_win_el = row.find("td", {"data-stat": "pts_win"})
        pts_lose_el = row.find("td", {"data-stat": "pts_lose"})

        if not winner_el or not loser_el:
            continue

        week = week_el.get_text(strip=True) if week_el else ""
        day = day_el.get_text(strip=True) if day_el else ""
        game_time = time_el.get_text(strip=True) if time_el else ""
        winner = winner_el.get_text(strip=True) if winner_el else ""
        loser = loser_el.get_text(strip=True) if loser_el else ""
        location = location_el.get_text(strip=True) if location_el else ""

        if date_el:
            csk = date_el.get("csk", "")
            text_date = date_el.get_text(strip=True)
            if csk and not csk.startswith("zz") and re.match(r'\d{4}-\d{2}-\d{2}', csk):
                date = csk
            else:
                date = text_date
        else:
            date = ""

        try:
            pts_win = int(pts_win_el.get_text(strip=True)) if pts_win_el else 0
        except (ValueError, AttributeError):
            pts_win = 0
        try:
            pts_lose = int(pts_lose_el.get_text(strip=True)) if pts_lose_el else 0
        except (ValueError, AttributeError):
            pts_lose = 0

        if not winner or winner.startswith("Week") or not any(c.isalpha() for c in winner):
            continue

        if location == "@":
            home_team = loser
            away_team = winner
            home_score = pts_lose
            away_score = pts_win
        else:
            home_team = winner
            away_team = loser
            home_score = pts_win
            away_score = pts_lose

        if pts_win == pts_lose:
            game_winner = "Tie"
        else:
            game_winner = winner

        score_diff = abs(pts_win - pts_lose)
        primetime = detect_primetime(day, game_time, year)

        games.append({
            "Season": year,
            "Week": week,
            "Day": day,
            "Date": date,
            "Time (ET)": game_time,
            "Home Team": home_team,
            "Away Team": away_team,
            "Home Score": home_score,
            "Away Score": away_score,
            "Score Difference": score_diff,
            "Winner": game_winner,
            "Primetime Slot": primetime,
        })

    return games


def write_excel(all_games, filepath):
    """Write all game data to a formatted Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NFL Scores"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="013369", end_color="013369", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    mnf_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    snf_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    tnf_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    for row_idx, game in enumerate(all_games, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = game.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_name not in ("Home Team", "Away Team", "Winner") else "left")

        primetime = game.get("Primetime Slot", "")
        if primetime:
            fill = None
            if "MNF" in primetime:
                fill = mnf_fill
            elif "SNF" in primetime:
                fill = snf_fill
            elif "TNF" in primetime:
                fill = tnf_fill
            if fill:
                ws.cell(row=row_idx, column=COLUMNS.index("Primetime Slot") + 1).fill = fill

    col_widths = {
        "Season": 9, "Week": 14, "Day": 6, "Date": 12, "Time (ET)": 10,
        "Home Team": 26, "Away Team": 26, "Home Score": 11, "Away Score": 11,
        "Score Difference": 14, "Winner": 26, "Primetime Slot": 18,
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(all_games) + 1}"
    wb.save(filepath)
    print(f"\nSaved to: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Scrape NFL scores from Pro Football Reference")
    parser.add_argument("--start-year", type=int, default=DEFAULT_START_YEAR, help="First season to scrape")
    parser.add_argument("--end-year", type=int, default=DEFAULT_END_YEAR, help="Last season to scrape")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file path")
    args = parser.parse_args()

    print("=" * 60)
    print("NFL Historical Scores Scraper")
    print(f"Seasons: {args.start_year} - {args.end_year}")
    print(f"Source: Pro Football Reference")
    print(f"Output: {args.output}")
    print("=" * 60)

    all_games = []
    failed_seasons = []

    for year in range(args.start_year, args.end_year + 1):
        print(f"\nScraping {year} season... ", end="", flush=True)
        games = scrape_season(year)

        if games:
            all_games.extend(games)
            print(f"{len(games)} games found")
        else:
            failed_seasons.append(year)
            print("NO GAMES FOUND")

        if year < args.end_year:
            time.sleep(REQUEST_DELAY)

    print("\n" + "=" * 60)
    print(f"Total games scraped: {len(all_games):,}")
    print(f"Seasons processed: {args.end_year - args.start_year + 1 - len(failed_seasons)}/{args.end_year - args.start_year + 1}")

    if failed_seasons:
        print(f"Failed seasons: {failed_seasons}")

    mnf = sum(1 for g in all_games if g["Primetime Slot"] == "MNF")
    snf = sum(1 for g in all_games if g["Primetime Slot"] == "SNF")
    tnf = sum(1 for g in all_games if g["Primetime Slot"] == "TNF")
    print(f"MNF games: {mnf:,} | SNF games: {snf:,} | TNF games: {tnf:,}")

    if all_games:
        print(f"\nWriting Excel file...")
        write_excel(all_games, args.output)
        file_size = os.path.getsize(args.output) / (1024 * 1024)
        print(f"File size: {file_size:.1f} MB")
    else:
        print("\nNo games to write!")

    print("\nDone!")


if __name__ == "__main__":
    main()
