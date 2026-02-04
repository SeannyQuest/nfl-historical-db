#!/usr/bin/env python3
"""
Download historical NFL betting spreads from Aussportsbetting.com
and merge them into the existing NFL score spreadsheets.

Adds 4 new columns:
  - Spread (Home Line Close)
  - Over/Under (Total Score Close)
  - Spread Result (Covered / Lost / Push)
  - O/U Result (Over / Under / Push)
"""

import argparse
import os
import subprocess
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BETTING_URL = "https://www.aussportsbetting.com/historical_data/nfl.xlsx"

# Columns in existing spreadsheet (0-indexed)
COL_DATE = 3
COL_HOME = 5
COL_AWAY = 6
COL_HOME_SCORE = 7
COL_AWAY_SCORE = 8

# New columns to add (1-indexed for openpyxl)
NEW_COL_SPREAD = 13
NEW_COL_OU = 14
NEW_COL_SPREAD_RESULT = 15
NEW_COL_OU_RESULT = 16

NEW_HEADERS = ["Spread", "Over/Under", "Spread Result", "O/U Result"]

# Aussportsbetting column indices (0-indexed)
AB_DATE = 0
AB_HOME = 1
AB_HOME_LINE_CLOSE = 19
AB_TOTAL_SCORE_CLOSE = 35


def download_betting_data():
    """Download the Aussportsbetting Excel file."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    print(f"Downloading betting data from {BETTING_URL}...")
    result = subprocess.run(
        ["curl", "-s", "-L", "-o", tmp.name, BETTING_URL],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print(f"  Download failed: {result.stderr}")
        return None
    size_kb = os.path.getsize(tmp.name) / 1024
    print(f"  Downloaded {size_kb:.0f} KB")
    return tmp.name


def build_betting_lookup(filepath):
    """Build a lookup dict: (date_str, home_team) -> {spread, ou}"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    lookup = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[AB_DATE] is None:
            continue

        raw_date = row[AB_DATE]
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            date_str = str(raw_date)[:10]

        home = str(row[AB_HOME] or "").strip()
        spread = row[AB_HOME_LINE_CLOSE]
        ou = row[AB_TOTAL_SCORE_CLOSE]

        if not home or not date_str:
            continue

        try:
            spread = float(spread) if spread is not None else None
        except (ValueError, TypeError):
            spread = None
        try:
            ou = float(ou) if ou is not None else None
        except (ValueError, TypeError):
            ou = None

        lookup[(date_str, home)] = {"spread": spread, "ou": ou}

    wb.close()
    print(f"  Built lookup with {len(lookup)} games")
    return lookup


def compute_spread_result(home_score, away_score, spread):
    if spread is None or home_score is None or away_score is None:
        return ""
    try:
        adjusted = home_score + spread
        if adjusted > away_score:
            return "Covered"
        elif adjusted < away_score:
            return "Lost"
        else:
            return "Push"
    except (ValueError, TypeError):
        return ""


def compute_ou_result(home_score, away_score, ou):
    if ou is None or home_score is None or away_score is None:
        return ""
    try:
        total = home_score + away_score
        if total > ou:
            return "Over"
        elif total < ou:
            return "Under"
        else:
            return "Push"
    except (ValueError, TypeError):
        return ""


def merge_into_excel(filepath, lookup):
    """Add spread columns to an existing NFL scores Excel file."""
    if not os.path.exists(filepath):
        print(f"  File not found: {filepath}")
        return 0, 0

    print(f"\nProcessing: {os.path.basename(filepath)}")
    wb = load_workbook(filepath)
    ws = wb.active

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="013369", end_color="013369", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    existing_header = ws.cell(row=1, column=NEW_COL_SPREAD).value
    if existing_header == "Spread":
        print("  Spread columns already exist — overwriting values")
    else:
        for i, header in enumerate(NEW_HEADERS):
            cell = ws.cell(row=1, column=NEW_COL_SPREAD + i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    covered_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    lost_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    over_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    matched = 0
    total = 0

    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=COL_DATE + 1).value
        home_val = ws.cell(row=row_idx, column=COL_HOME + 1).value
        home_score = ws.cell(row=row_idx, column=COL_HOME_SCORE + 1).value
        away_score = ws.cell(row=row_idx, column=COL_AWAY_SCORE + 1).value

        if date_val is None or home_val is None:
            continue

        total += 1
        date_str = str(date_val).strip()[:10]
        home_str = str(home_val).strip()

        betting = lookup.get((date_str, home_str))

        if betting:
            matched += 1
            spread = betting["spread"]
            ou = betting["ou"]

            try:
                hs = int(home_score) if home_score is not None else None
            except (ValueError, TypeError):
                hs = None
            try:
                as_ = int(away_score) if away_score is not None else None
            except (ValueError, TypeError):
                as_ = None

            spread_result = compute_spread_result(hs, as_, spread)
            ou_result = compute_ou_result(hs, as_, ou)

            ws.cell(row=row_idx, column=NEW_COL_SPREAD, value=spread)
            ws.cell(row=row_idx, column=NEW_COL_OU, value=ou)

            sr_cell = ws.cell(row=row_idx, column=NEW_COL_SPREAD_RESULT, value=spread_result)
            ou_cell = ws.cell(row=row_idx, column=NEW_COL_OU_RESULT, value=ou_result)

            for c in [ws.cell(row=row_idx, column=NEW_COL_SPREAD),
                       ws.cell(row=row_idx, column=NEW_COL_OU),
                       sr_cell, ou_cell]:
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border

            if spread_result == "Covered":
                sr_cell.fill = covered_fill
            elif spread_result == "Lost":
                sr_cell.fill = lost_fill
            if ou_result == "Over":
                ou_cell.fill = over_fill
        else:
            for col in range(NEW_COL_SPREAD, NEW_COL_OU_RESULT + 1):
                cell = ws.cell(row=row_idx, column=col, value=None)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

    col_widths = {NEW_COL_SPREAD: 10, NEW_COL_OU: 12, NEW_COL_SPREAD_RESULT: 14, NEW_COL_OU_RESULT: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(NEW_COL_OU_RESULT)}{ws.max_row}"

    wb.save(filepath)
    print(f"  Matched: {matched}/{total} games ({matched/total*100:.1f}%)" if total > 0 else "  No games found")
    return matched, total


def main():
    parser = argparse.ArgumentParser(description="Merge betting spreads into NFL Excel files")
    parser.add_argument("--files", nargs="+", help="Excel files to process")
    args = parser.parse_args()

    excel_files = args.files if args.files else []
    if not excel_files:
        print("No files specified. Use --files <path1> <path2> ...")
        return

    print("=" * 60)
    print("NFL Betting Spreads Merger")
    print(f"Files: {', '.join(os.path.basename(f) for f in excel_files)}")
    print("=" * 60)

    betting_file = download_betting_data()
    if not betting_file:
        print("Failed to download betting data!")
        return

    try:
        print("\nBuilding betting data lookup...")
        lookup = build_betting_lookup(betting_file)

        total_matched = 0
        total_games = 0
        for filepath in excel_files:
            m, t = merge_into_excel(filepath, lookup)
            total_matched += m
            total_games += t

        print("\n" + "=" * 60)
        print(f"Total: {total_matched} games matched with betting data out of {total_games}")
        print("=" * 60)
    finally:
        if betting_file and os.path.exists(betting_file):
            os.unlink(betting_file)

    print("\nDone!")


if __name__ == "__main__":
    main()
