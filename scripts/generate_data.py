#!/usr/bin/env python3
"""Convert NFL Excel files into a compact data.js for the web app."""

import argparse
import os
import json
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUTPUT = os.path.join(SCRIPT_DIR, "..", "public", "data.js")

# Column indices (1-based from Excel)
COL = {
    "Season": 1, "Week": 2, "Day": 3, "Date": 4, "Time": 5,
    "Home Team": 6, "Away Team": 7, "Home Score": 8, "Away Score": 9,
    "Score Diff": 10, "Winner": 11, "Primetime": 12,
    "Spread": 13, "Over/Under": 14, "Spread Result": 15, "O/U Result": 16,
}


def read_excel(filepath):
    """Read an Excel file and return list of game dicts."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    games = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]
        if len(cells) < 12 or cells[0] is None:
            continue

        season = cells[COL["Season"] - 1]
        week = str(cells[COL["Week"] - 1] or "")
        day = str(cells[COL["Day"] - 1] or "")
        date = str(cells[COL["Date"] - 1] or "")
        time_et = str(cells[COL["Time"] - 1] or "")
        home = str(cells[COL["Home Team"] - 1] or "")
        away = str(cells[COL["Away Team"] - 1] or "")
        hs = cells[COL["Home Score"] - 1]
        away_s = cells[COL["Away Score"] - 1]
        pt = str(cells[COL["Primetime"] - 1] or "")

        if not home or not away:
            continue
        if pt == "None":
            pt = ""
        if time_et == "None":
            time_et = ""

        try:
            hs = int(hs) if hs is not None else 0
        except (ValueError, TypeError):
            hs = 0
        try:
            away_s = int(away_s) if away_s is not None else 0
        except (ValueError, TypeError):
            away_s = 0

        # Read spread/betting columns (may not exist in older files)
        spread = cells[COL["Spread"] - 1] if len(cells) >= 13 else None
        ou = cells[COL["Over/Under"] - 1] if len(cells) >= 14 else None
        spread_result = cells[COL["Spread Result"] - 1] if len(cells) >= 15 else None
        ou_result = cells[COL["O/U Result"] - 1] if len(cells) >= 16 else None

        try:
            spread = round(float(spread), 1) if spread is not None else None
        except (ValueError, TypeError):
            spread = None
        try:
            ou = round(float(ou), 1) if ou is not None else None
        except (ValueError, TypeError):
            ou = None

        spread_result = str(spread_result or "")
        ou_result = str(ou_result or "")
        if spread_result == "None":
            spread_result = ""
        if ou_result == "None":
            ou_result = ""

        game = {
            "s": int(season),
            "w": week,
            "d": day,
            "dt": date,
            "tm": time_et,
            "h": home,
            "a": away,
            "hs": hs,
            "as": away_s,
            "pt": pt,
        }
        if spread is not None:
            game["sp"] = spread
        if ou is not None:
            game["ou"] = ou
        if spread_result:
            game["sr"] = spread_result
        if ou_result:
            game["our"] = ou_result

        games.append(game)

    wb.close()
    return games


FRANCHISE_MAP = {
    "Indianapolis Colts": "Colts",
    "Baltimore Colts": "Colts",
    "Las Vegas Raiders": "Raiders",
    "Oakland Raiders": "Raiders",
    "Los Angeles Raiders": "Raiders",
    "Los Angeles Chargers": "Chargers",
    "San Diego Chargers": "Chargers",
    "Los Angeles Rams": "Rams",
    "St. Louis Rams": "Rams",
    "Cleveland Rams": "Rams",
    "Tennessee Titans": "Titans",
    "Tennessee Oilers": "Titans",
    "Houston Oilers": "Titans",
    "Arizona Cardinals": "Cardinals",
    "Phoenix Cardinals": "Cardinals",
    "St. Louis Cardinals": "Cardinals",
    "Chicago Cardinals": "Cardinals",
    "Washington Commanders": "Washington",
    "Washington Football Team": "Washington",
    "Washington Redskins": "Washington",
    "New England Patriots": "Patriots",
    "Boston Patriots": "Patriots",
    "Houston Texans": "Texans",
}


def main():
    parser = argparse.ArgumentParser(description="Generate data.js from NFL Excel files")
    parser.add_argument("--files", nargs="+", help="Excel files to read")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output data.js path")
    args = parser.parse_args()

    if not args.files:
        print("No files specified. Use --files <path1> <path2> ...")
        return

    all_games = []
    for filepath in args.files:
        if not os.path.exists(filepath):
            print(f"Warning: {filepath} not found, skipping")
            continue
        games = read_excel(filepath)
        print(f"Read {len(games)} games from {os.path.basename(filepath)}")
        all_games.extend(games)

    # Deduplicate by (season, date, home, away)
    seen = set()
    unique_games = []
    for g in all_games:
        key = (g["s"], g["dt"], g["h"], g["a"])
        if key not in seen:
            seen.add(key)
            unique_games.append(g)
    print(f"Total unique games: {len(unique_games)}")

    # Sort by date
    unique_games.sort(key=lambda g: (g["s"], g["dt"]))

    # Build JS output
    lines = []
    for g in unique_games:
        js = json.dumps(g, separators=(',', ':'))
        lines.append(js)

    franchise_js = json.dumps(FRANCHISE_MAP, indent=2)

    js_content = f"// Auto-generated from NFL Excel data — {len(unique_games)} games\n"
    js_content += f"const NFL_GAMES = [\n" + ",\n".join(lines) + "\n];\n\n"
    js_content += f"const FRANCHISE_MAP = {franchise_js};\n"

    # Ensure output directory exists
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)

    with open(args.output, "w") as f:
        f.write(js_content)

    size_kb = os.path.getsize(args.output) / 1024
    print(f"Written {args.output} ({size_kb:.0f} KB)")


if __name__ == "__main__":
    main()
